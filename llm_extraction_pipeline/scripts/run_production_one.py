"""
Process one paper through the LLM extraction pipeline and push to GitHub.

Usage:
    python run_production_one.py --index 1
    python run_production_one.py --index 1 --no-push
    python run_production_one.py --status

Environment variables required:
    OPENAI_API_KEY
    MODEL_EXTRACT   (default: gpt-4.1)

Output: llm_run/production/CBM0001/
    context_packet.json
    raw_response.json
    extraction.json
    CBM0001_<doi>_llm_extraction.xlsx
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
import urllib.error
import urllib.request
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


ROOT = Path(os.environ.get(
    "CEMENT_DB_ROOT",
    "/Users/yanjie/Library/Mobile Documents/com~apple~CloudDocs/Academic/computing/260622 Cement database Journal",
))
MANIFEST_PATH = ROOT / "llm_extraction_pipeline/manifest.json"
SCHEMA_PATH = ROOT / "llm_extraction_pipeline/schemas/cement_extraction.schema.json"
PROMPT_PATH = ROOT / "llm_extraction_pipeline/prompts/extract_paper.md"
PRODUCTION_DIR = ROOT / "llm_run/production"


def load_manifest() -> list[dict]:
    return json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))


def clean(text: str) -> str:
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text or "")


def compact(text: str, limit: int = 3500) -> str:
    text = re.sub(r"\s+", " ", clean(text)).strip()
    return text[:limit]


def prepare_packet(entry: dict) -> dict:
    pdf_path = ROOT / "cbm_papers_subset_200" / entry["pdf"]
    paper_id = entry["paper_id"]

    reader = PdfReader(str(pdf_path))
    pages = []
    for i, page in enumerate(reader.pages, 1):
        try:
            text = clean(page.extract_text() or "")
        except Exception:
            text = ""
        score = 0
        reasons = []
        if i == 1:
            score += 3
            reasons.append("first page")
        if re.search(r"compressive strength|compressive stress|cube strength|cylinder strength|strength tests", text, re.I):
            score += 4
            reasons.append("compressive-strength keyword")
        if re.search(r"mix|mixture|proportion|cement|water|aggregate|w/c|water[- ]to[- ]cement", text, re.I):
            score += 2
            reasons.append("mixture/material keyword")
        if re.search(r"cur(?:ed|ing)|specimen|compressive.*test|test(?:ed|ing)", text, re.I):
            score += 1
            reasons.append("methods/conditions keyword")
        if score:
            pages.append({"page": i, "score": score, "reason": reasons, "text": compact(text)})

    pages = sorted(pages, key=lambda x: (-x["score"], x["page"]))[:8]
    keep_page_nums = {p["page"] for p in pages}

    tables = []
    with pdfplumber.open(str(pdf_path)) as doc:
        for page_num in sorted(keep_page_nums):
            if page_num < 1 or page_num > len(doc.pages):
                continue
            page = doc.pages[page_num - 1]
            try:
                page_tables = page.extract_tables() or []
            except Exception:
                page_tables = []
            for ti, table in enumerate(page_tables, 1):
                rows = []
                for row in table:
                    vals = [compact(str(v or ""), 500) for v in row]
                    if any(vals):
                        rows.append(vals)
                if not rows:
                    continue
                text = "\n".join(" | ".join(r) for r in rows)
                low = text.lower()
                if not any(k in low for k in ["compressive", "strength", "cement", "water", "aggregate", "mix", "proportion"]):
                    continue
                tables.append({
                    "table_id": f"T{page_num:03d}_{ti:02d}",
                    "page": page_num,
                    "caption": "",
                    "text": compact(text, 3000),
                    "csv_path": "",
                })

    figures = []
    for page in pages:
        if re.search(r"fig(?:ure)?\.?\s*\d+", page["text"], re.I) and re.search(r"compressive strength|strength tests", page["text"], re.I):
            figures.append({
                "figure_id": f"F_PAGE_{page['page']:03d}",
                "page": page["page"],
                "caption": "candidate compressive-strength figure/page",
                "image_path": "",
                "reason": ["figure/page mentions compressive strength"],
            })

    return {
        "paper": {
            "paper_id": paper_id,
            "pdf_filename": entry["pdf"],
            "doi": entry["doi"],
            "title": "",
        },
        "pages": pages,
        "tables": tables[:10],
        "figures": figures[:6],
    }


def build_prompt(packet: dict) -> str:
    prompt = PROMPT_PATH.read_text(encoding="utf-8")
    selected_pages = "\n\n".join(
        f"PAGE {p['page']} | reason={p['reason']}\n{p['text']}" for p in packet["pages"]
    )
    tables = "\n\n".join(
        f"{t['table_id']} | page={t['page']}\n{t['text']}" for t in packet["tables"]
    )
    figures = "\n\n".join(
        f"{f['figure_id']} | page={f['page']} | caption={f['caption']} | reason={f['reason']}"
        for f in packet["figures"]
    )
    return (
        prompt
        .replace("{{paper_metadata_json}}", json.dumps(packet["paper"], ensure_ascii=False, indent=2))
        .replace("{{selected_page_text}}", selected_pages)
        .replace("{{candidate_tables_text}}", tables)
        .replace("{{figure_packet_text}}", figures)
        .replace("{{few_shot_examples_json}}", "[]")
    )


def call_api(prompt: str) -> dict:
    key = os.environ.get("OPENAI_API_KEY")
    if not key:
        raise SystemExit("OPENAI_API_KEY is not set")
    model = os.environ.get("MODEL_EXTRACT", "gpt-4.1")
    base_url = (os.environ.get("OPENAI_BASE_URL") or os.environ.get("API_BASE_URL") or "https://api.openai.com/v1").rstrip("/")
    schema = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    body = {
        "model": model,
        "input": [{"role": "user", "content": [{"type": "input_text", "text": prompt}]}],
        "text": {
            "format": {
                "type": "json_schema",
                "name": "cement_extraction",
                "schema": schema,
                "strict": True,
            }
        },
    }
    req = urllib.request.Request(
        f"{base_url}/responses",
        data=json.dumps(body).encode("utf-8"),
        headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=300) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        error_text = e.read().decode("utf-8", errors="replace")
        raise SystemExit(f"HTTPError {e.code}: {error_text[:2000]}")


def response_text(raw: dict) -> str:
    if isinstance(raw.get("output_text"), str):
        return raw["output_text"]
    chunks = []
    for item in raw.get("output", []):
        for content in item.get("content", []):
            if content.get("type") in {"output_text", "text"} and isinstance(content.get("text"), str):
                chunks.append(content["text"])
    return "\n".join(chunks)


def add_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([
            ";".join(str(x) for x in row.get(h, [])) if isinstance(row.get(h), list) else row.get(h, "")
            for h in headers
        ])
    color_map = {"README": "5B2C6F", "review_queue": "7F1D1D", "table_preview": "3D5A2A"}
    fill = PatternFill("solid", fgColor=color_map.get(name, "1F4E79"))
    header_font = Font(name="Aptos", bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=10)
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = min(max(12, len(str(ws.cell(1, col).value or "")) + 2), 36)


def build_workbook(data: dict, entry: dict, run_dir: Path) -> Path:
    paper_id = entry["paper_id"]
    doi_slug = entry["doi"].replace("/", "_").replace(".", "_")
    wb_path = run_dir / f"{paper_id}_{doi_slug}_llm_extraction.xlsx"

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "README", ["Item", "Value"], [
        {"Item": "Paper ID", "Value": paper_id},
        {"Item": "DOI", "Value": entry["doi"]},
        {"Item": "PDF", "Value": entry["pdf"]},
        {"Item": "Model", "Value": os.environ.get("MODEL_EXTRACT", "gpt-4.1")},
    ])
    paper = data.get("paper", {})
    add_sheet(wb, "paper", list(paper.keys()) or ["paper_id"], [paper])
    sheet_headers = {
        "mixture_components": ["component_id", "paper_id", "mixture_id", "mixture_original_id", "mixture_description", "component_original", "component_standard", "category", "role_in_mix", "amount", "unit", "basis", "replacement_target", "replacement_pct", "reported_w_c_ratio", "reported_w_p_ratio", "evidence_ids", "confidence", "human_decision", "corrected_amount", "notes"],
        "strength_results": ["result_id", "paper_id", "mixture_id", "mixture_original_id", "property", "age_days", "unit", "table_value", "recommended_value", "final_value", "value_status", "is_approximate", "estimated_error_mpa", "curing_condition_id", "testing_condition_id", "evidence_ids", "confidence", "human_decision", "notes"],
        "conditions": ["condition_id", "paper_id", "condition_type", "applies_to", "specimen_type", "temperature_c", "rh_pct", "curing_method", "demould_time_days", "test_age_days", "test_method", "test_standard", "specimen_geometry", "exposure_solution", "exposure_duration_days", "evidence_ids", "confidence", "human_decision", "notes"],
        "figures": ["figure_id", "paper_id", "page", "caption", "figure_type", "strength_relevance", "image_path", "extraction_status", "confidence", "human_decision", "notes"],
        "evidence": ["evidence_id", "paper_id", "source_modality", "page", "source_location", "evidence_text", "supports", "confidence", "human_decision", "notes"],
        "review_queue": ["review_id", "paper_id", "target_sheet", "target_record_ids", "priority", "question", "suggested_action", "human_answer", "final_decision", "notes"],
    }
    for sheet, headers in sheet_headers.items():
        add_sheet(wb, sheet, headers, data.get(sheet, []))
    add_sheet(wb, "data_dictionary", ["field_or_concept", "definition"], [
        {"field_or_concept": "confidence", "definition": "LLM confidence before human confirmation"},
        {"field_or_concept": "review_queue", "definition": "Ambiguous or figure-derived items for human review"},
        {"field_or_concept": "value_status", "definition": "exact_table_value | explicit_text_value | figure_derived | inferred"},
    ])
    add_sheet(wb, "table_preview", ["source", "page", "score", "text_preview", "csv_path"], [])
    wb.save(str(wb_path))
    return wb_path


def git_commit_push(paper_id: str, doi: str, run_dir: Path, wb_path: Path) -> None:
    rel_run = str(run_dir.relative_to(ROOT))
    subprocess.run(["git", "-C", str(ROOT), "add", rel_run], check=True)
    msg = f"LLM extraction: {paper_id} ({doi})"
    subprocess.run(["git", "-C", str(ROOT), "commit", "-m", msg], check=True)
    subprocess.run(["git", "-C", str(ROOT), "push"], check=True)
    print(f"  Pushed: {msg}")


def show_status() -> None:
    manifest = load_manifest()
    done = []
    todo = []
    for entry in manifest:
        run_dir = PRODUCTION_DIR / entry["paper_id"]
        if (run_dir / "extraction.json").exists():
            done.append(entry["paper_id"])
        else:
            todo.append(entry["paper_id"])
    print(f"Done:  {len(done)}/200")
    print(f"Todo:  {len(todo)}/200")
    if todo:
        print(f"Next:  {todo[0]}  (index {manifest[int(todo[0][3:]) - 1]['index']})")


def process_paper(index: int, push: bool = True) -> None:
    manifest = load_manifest()
    if index < 1 or index > len(manifest):
        raise SystemExit(f"Index must be 1-{len(manifest)}")

    entry = manifest[index - 1]
    paper_id = entry["paper_id"]
    run_dir = PRODUCTION_DIR / paper_id

    if (run_dir / "extraction.json").exists():
        print(f"SKIP {paper_id}: already extracted")
        return

    print(f"Processing {paper_id}  ({entry['doi']})")
    run_dir.mkdir(parents=True, exist_ok=True)

    print("  Preparing packet...")
    packet = prepare_packet(entry)
    (run_dir / "context_packet.json").write_text(json.dumps(packet, ensure_ascii=False, indent=2), encoding="utf-8")

    print("  Calling API...")
    prompt = build_prompt(packet)
    raw = call_api(prompt)
    (run_dir / "raw_response.json").write_text(json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8")

    text = response_text(raw)
    if not text:
        raise SystemExit("API response contained no output text")
    data = json.loads(text)
    (run_dir / "extraction.json").write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print("  Building workbook...")
    wb_path = build_workbook(data, entry, run_dir)

    result = {
        "paper_id": paper_id,
        "doi": entry["doi"],
        "strength_results": len(data.get("strength_results", [])),
        "direct_candidates": sum(1 for r in data.get("strength_results", []) if r.get("confidence", 0) >= 0.9),
        "review_queue": len(data.get("review_queue", [])),
        "workbook": wb_path.name,
    }
    print(f"  Done: {result}")

    if push:
        print("  Committing and pushing...")
        git_commit_push(paper_id, entry["doi"], run_dir, wb_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Run LLM extraction for one paper")
    parser.add_argument("--index", type=int, help="Paper index 1-200")
    parser.add_argument("--no-push", action="store_true", help="Skip git push")
    parser.add_argument("--status", action="store_true", help="Show extraction status")
    args = parser.parse_args()

    if args.status:
        show_status()
        return

    if args.index is None:
        parser.print_help()
        sys.exit(1)

    process_paper(args.index, push=not args.no_push)


if __name__ == "__main__":
    main()
