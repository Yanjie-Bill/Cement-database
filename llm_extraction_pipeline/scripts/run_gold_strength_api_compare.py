from __future__ import annotations

import json
import os
import re
import urllib.error
import urllib.request
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/yanjie/Library/Mobile Documents/com~apple~CloudDocs/Academic/computing/260622 Cement database Journal")
PDF = ROOT / "cbm_papers_subset_200/10.1016_j.conbuildmat.2009.12.024.pdf"
GOLD = ROOT / "annotation/CBM0008_10.1016_j.conbuildmat.2009.12.024_clean_annotation_for_review.xlsx"
RUN_DIR = ROOT / "llm_run/api_gold_test_CBM0008_strength"
RAW_PATH = RUN_DIR / "raw_response.json"
PRED_JSON = RUN_DIR / "strength_prediction.json"
PRED_XLSX = RUN_DIR / "CBM0008_strength_api_prediction.xlsx"
COMPARE_CSV = RUN_DIR / "strength_comparison.csv"
REPORT_MD = RUN_DIR / "comparison_report.md"


def get_table3_text() -> str:
    with pdfplumber.open(PDF) as doc:
        page = doc.pages[4]
        text = page.extract_text() or ""
        tables = page.extract_tables() or []
    table_text = ""
    for table in tables:
        rows = []
        for row in table:
            vals = [re.sub(r"\s+", " ", str(v or "")).strip() for v in row]
            if any(vals):
                rows.append(" | ".join(vals))
        joined = "\n".join(rows)
        if "Compressivestrength" in joined.replace(" ", "") or "M2012" in joined:
            table_text = joined
            break
    if not table_text:
        table_text = text
    return table_text


def call_api(table_text: str) -> dict:
    key = os.environ.get("OPENAI_API_KEY")
    if not key:
        raise SystemExit("OPENAI_API_KEY is not set")
    model = os.environ.get("MODEL_EXTRACT", "gpt-5.5")
    base_url = (os.environ.get("OPENAI_BASE_URL") or "https://api.302.ai/v1").rstrip("/")
    schema = {
        "type": "object",
        "additionalProperties": False,
        "required": ["paper_id", "source", "strength_results", "warnings"],
        "properties": {
            "paper_id": {"type": "string"},
            "source": {"type": "string"},
            "warnings": {"type": "array", "items": {"type": "string"}},
            "strength_results": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "required": ["mixture_original_id", "age_days", "value_mpa", "std_dev_mpa", "unit", "confidence", "evidence"],
                    "properties": {
                        "mixture_original_id": {"type": "string"},
                        "age_days": {"type": "number"},
                        "value_mpa": {"type": "number"},
                        "std_dev_mpa": {"type": ["number", "null"]},
                        "unit": {"type": "string"},
                        "confidence": {"type": "number", "minimum": 0, "maximum": 1},
                        "evidence": {"type": "string"}
                    }
                }
            }
        }
    }
    prompt = f"""
You are extracting a gold-test table for a cement/concrete compressive-strength database.

Extract ONLY compressive strength from Table 3 below.

Rules:
- Paper ID must be CBM0008.
- The table has curing ages 1, 4, 7, 14, 28, 90, 180 days.
- For each mixture row, each age has a pair: compressive strength mean, then standard deviation r.
- Skip rows where all values are dash/missing, e.g. failed formulations.
- Return one strength_results record per mixture-age mean value.
- Use unit MPa.
- Do not extract flexural strength or any other property.
- Expected count is 112 if all non-missing rows are extracted.

Table 3 text:
{table_text}
""".strip()
    body = {
        "model": model,
        "input": [{"role": "user", "content": [{"type": "input_text", "text": prompt}]}],
        "text": {"format": {"type": "json_schema", "name": "cbm0008_strength_table", "schema": schema, "strict": True}},
    }
    req = urllib.request.Request(
        f"{base_url}/responses",
        data=json.dumps(body).encode("utf-8"),
        headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=360) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        error_text = e.read().decode("utf-8", errors="replace")
        RAW_PATH.write_text(error_text, encoding="utf-8")
        raise SystemExit(f"HTTPError {e.code}: {error_text[:2000]}")
    RAW_PATH.write_text(json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8")
    chunks = []
    for item in raw.get("output", []):
        for content in item.get("content", []):
            if isinstance(content.get("text"), str):
                chunks.append(content["text"])
    text = raw.get("output_text") or "\n".join(chunks)
    if not text:
        raise SystemExit("No output_text in API response")
    data = json.loads(text)
    PRED_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data


def read_gold() -> list[dict]:
    wb = load_workbook(GOLD, data_only=True, read_only=True)
    ws = wb["strength_results"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(x or "") for x in rows[0]]
    out = []
    for row in rows[1:]:
        rec = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        if rec.get("property") != "compressive_strength":
            continue
        val = rec.get("final_value") or rec.get("recommended_value") or rec.get("table_value")
        if val in (None, ""):
            continue
        out.append({
            "mixture_original_id": str(rec.get("mixture_original_id") or rec.get("mixture_id")),
            "age_days": float(rec.get("age_days")),
            "value_mpa": float(val),
            "std_dev_mpa": None if rec.get("std_dev_mpa") in (None, "") else float(rec.get("std_dev_mpa")),
        })
    return out


def compare(pred: list[dict], gold: list[dict]) -> tuple[list[dict], dict]:
    gold_key = {(g["mixture_original_id"], float(g["age_days"])): g for g in gold}
    pred_key = {(p["mixture_original_id"], float(p["age_days"])): p for p in pred}
    rows = []
    tp = fp = fn = value_exact = std_exact = 0
    for key in sorted(set(gold_key) | set(pred_key)):
        g = gold_key.get(key)
        p = pred_key.get(key)
        status = ""
        diff = ""
        std_diff = ""
        if g and p:
            tp += 1
            diff = abs(float(p["value_mpa"]) - float(g["value_mpa"]))
            if diff <= 1e-9:
                value_exact += 1
            if g.get("std_dev_mpa") is not None and p.get("std_dev_mpa") is not None:
                std_diff = abs(float(p["std_dev_mpa"]) - float(g["std_dev_mpa"]))
                if std_diff <= 1e-9:
                    std_exact += 1
            status = "matched_key"
        elif p and not g:
            fp += 1
            status = "false_positive"
        else:
            fn += 1
            status = "false_negative"
        rows.append({
            "mixture_original_id": key[0],
            "age_days": key[1],
            "gold_value_mpa": "" if not g else g["value_mpa"],
            "pred_value_mpa": "" if not p else p["value_mpa"],
            "value_abs_diff": diff,
            "gold_std_dev_mpa": "" if not g else g.get("std_dev_mpa"),
            "pred_std_dev_mpa": "" if not p else p.get("std_dev_mpa"),
            "std_abs_diff": std_diff,
            "status": status,
        })
    precision = tp / (tp + fp) if tp + fp else 0
    recall = tp / (tp + fn) if tp + fn else 0
    value_accuracy = value_exact / tp if tp else 0
    std_accuracy = std_exact / tp if tp else 0
    metrics = {
        "gold_records": len(gold),
        "pred_records": len(pred),
        "tp_by_mixture_age": tp,
        "fp_by_mixture_age": fp,
        "fn_by_mixture_age": fn,
        "precision_by_mixture_age": precision,
        "recall_by_mixture_age": recall,
        "exact_value_accuracy_among_matched": value_accuracy,
        "exact_std_accuracy_among_matched": std_accuracy,
    }
    return rows, metrics


def write_outputs(data: dict, rows: list[dict], metrics: dict) -> None:
    import csv
    with COMPARE_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "prediction"
    headers = ["mixture_original_id", "age_days", "value_mpa", "std_dev_mpa", "unit", "confidence", "evidence"]
    ws.append(headers)
    for r in data["strength_results"]:
        ws.append([r.get(h, "") for h in headers])
    ws2 = wb.create_sheet("comparison")
    ws2.append(list(rows[0].keys()))
    for r in rows:
        ws2.append([r[k] for k in rows[0].keys()])
    ws3 = wb.create_sheet("metrics")
    ws3.append(["metric", "value"])
    for k, v in metrics.items():
        ws3.append([k, v])
    for wsx in wb.worksheets:
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="E5E7EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in wsx[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
        wsx.freeze_panes = "A2"
        wsx.auto_filter.ref = wsx.dimensions
        for col in range(1, wsx.max_column + 1):
            wsx.column_dimensions[get_column_letter(col)].width = 18
        for row in wsx.iter_rows():
            for cell in row:
                cell.border = border
    wb.save(PRED_XLSX)
    REPORT_MD.write_text(
        "# CBM0008 API Strength Extraction Comparison\n\n"
        f"- Gold records: {metrics['gold_records']}\n"
        f"- Predicted records: {metrics['pred_records']}\n"
        f"- TP by mixture+age: {metrics['tp_by_mixture_age']}\n"
        f"- FP by mixture+age: {metrics['fp_by_mixture_age']}\n"
        f"- FN by mixture+age: {metrics['fn_by_mixture_age']}\n"
        f"- Precision by mixture+age: {metrics['precision_by_mixture_age']:.3f}\n"
        f"- Recall by mixture+age: {metrics['recall_by_mixture_age']:.3f}\n"
        f"- Exact value accuracy among matched keys: {metrics['exact_value_accuracy_among_matched']:.3f}\n"
        f"- Exact std-dev accuracy among matched keys: {metrics['exact_std_accuracy_among_matched']:.3f}\n",
        encoding="utf-8",
    )


def main() -> None:
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    table_text = get_table3_text()
    data = call_api(table_text)
    gold = read_gold()
    pred = data.get("strength_results", [])
    rows, metrics = compare(pred, gold)
    write_outputs(data, rows, metrics)
    print(json.dumps({
        "prediction_json": str(PRED_JSON.relative_to(ROOT)),
        "comparison_csv": str(COMPARE_CSV.relative_to(ROOT)),
        "comparison_workbook": str(PRED_XLSX.relative_to(ROOT)),
        "report": str(REPORT_MD.relative_to(ROOT)),
        **metrics,
    }, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
