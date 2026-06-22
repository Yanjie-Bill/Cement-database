#!/usr/bin/env python3
"""Import reviewed cement annotation workbooks into a local SQLite database."""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_ANNOTATION_DIR = ROOT / "annotation"
DEFAULT_DB = Path(__file__).resolve().parent / "cement_review.db"

ID_COLUMNS = {
    "paper": ["paper_id"],
    "mixture_components": ["component_id"],
    "strength_results": ["result_id"],
    "conditions": ["condition_id"],
    "figures": ["figure_id"],
    "evidence": ["evidence_id"],
    "review_queue": ["review_id", "queue_id", "id"],
}


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def norm_header(value: Any, index: int) -> str:
    text = str(value or "").strip()
    if not text:
        return f"column_{index}"
    text = re.sub(r"\s+", "_", text.lower())
    text = re.sub(r"[^a-z0-9_]+", "_", text).strip("_")
    return text or f"column_{index}"


def rows_from_sheet(ws) -> list[dict[str, Any]]:
    raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = []
    seen: dict[str, int] = {}
    for i, header in enumerate(raw_headers, 1):
        name = norm_header(header, i)
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        headers.append(name)

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = [clean_value(v) for v in row]
        if not any(v is not None for v in values):
            continue
        item = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        rows.append({k: v for k, v in item.items() if v is not None})
    return rows


def pick(row: dict[str, Any], names: list[str]) -> Any:
    for name in names:
        if row.get(name) is not None:
            return row[name]
    return None


def infer_paper_id(workbook: Path, paper_rows: list[dict[str, Any]]) -> str:
    value = pick(paper_rows[0], ["paper_id"]) if paper_rows else None
    if value:
        return str(value)
    match = re.match(r"(CBM\d+)", workbook.name)
    return match.group(1) if match else workbook.stem


def truthy_decision(value: Any) -> str:
    return str(value or "").strip().lower()


def initial_status(row: dict[str, Any]) -> str:
    answer = truthy_decision(pick(row, ["human_answer", "answer"]))
    final = truthy_decision(pick(row, ["final_decision", "human_decision", "decision"]))
    if (answer or final) and not any(token in final for token in ["review", "uncertain", "check", "needs", "flag"]):
        return "resolved"
    decision = truthy_decision(pick(row, ["human_decision", "decision", "status", "final_decision"]))
    if any(token in decision for token in ["review", "uncertain", "check", "needs", "flag"]):
        return "open"
    if any(token in decision for token in ["accept", "correct", "verified", "confirmed", "ok"]):
        return "resolved"
    return "open"


def connect(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS source_workbooks (
            filename TEXT PRIMARY KEY,
            path TEXT NOT NULL,
            imported_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS papers (
            paper_id TEXT PRIMARY KEY,
            source_workbook TEXT NOT NULL,
            title TEXT,
            doi TEXT,
            year TEXT,
            journal TEXT,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_workbook TEXT NOT NULL,
            paper_id TEXT NOT NULL,
            sheet TEXT NOT NULL,
            record_key TEXT NOT NULL,
            human_decision TEXT,
            confidence REAL,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(source_workbook, sheet, record_key)
        );

        CREATE TABLE IF NOT EXISTS review_queue (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_workbook TEXT NOT NULL,
            paper_id TEXT NOT NULL,
            review_key TEXT NOT NULL,
            target_sheet TEXT,
            target_record_ids TEXT,
            priority TEXT,
            question TEXT,
            suggested_action TEXT,
            status TEXT NOT NULL DEFAULT 'open',
            final_decision TEXT,
            human_answer TEXT,
            notes TEXT,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(source_workbook, review_key)
        );

        CREATE TABLE IF NOT EXISTS answers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            review_item_id INTEGER NOT NULL,
            participant TEXT NOT NULL,
            decision TEXT NOT NULL,
            answer TEXT,
            corrected_value TEXT,
            comment TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(review_item_id) REFERENCES review_queue(id)
        );

        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_type TEXT NOT NULL,
            payload_json TEXT NOT NULL,
            created_at TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_records_paper_sheet ON records(paper_id, sheet);
        CREATE INDEX IF NOT EXISTS idx_review_status ON review_queue(status);
        CREATE INDEX IF NOT EXISTS idx_answers_review ON answers(review_item_id);
        """
    )


def import_workbook(conn: sqlite3.Connection, workbook: Path) -> dict[str, int]:
    wb = load_workbook(workbook, data_only=True, read_only=True)
    imported_at = now_iso()
    sheet_rows: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in wb.sheetnames:
        normalized = norm_header(sheet_name, 1)
        if normalized in {"readme", "data_dictionary", "table_preview"}:
            continue
        sheet_rows[normalized] = rows_from_sheet(wb[sheet_name])

    paper_rows = sheet_rows.get("paper", [])
    paper_id = infer_paper_id(workbook, paper_rows)
    paper_data = paper_rows[0] if paper_rows else {"paper_id": paper_id}
    title = pick(paper_data, ["title", "paper_title"])
    doi = pick(paper_data, ["doi"])
    year = pick(paper_data, ["year", "publication_year"])
    journal = pick(paper_data, ["journal", "source"])

    conn.execute(
        """
        INSERT INTO source_workbooks(filename, path, imported_at)
        VALUES (?, ?, ?)
        ON CONFLICT(filename) DO UPDATE SET path=excluded.path, imported_at=excluded.imported_at
        """,
        (workbook.name, str(workbook), imported_at),
    )
    conn.execute(
        """
        INSERT INTO papers(paper_id, source_workbook, title, doi, year, journal, data_json, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(paper_id) DO UPDATE SET
            source_workbook=excluded.source_workbook,
            title=excluded.title,
            doi=excluded.doi,
            year=excluded.year,
            journal=excluded.journal,
            data_json=excluded.data_json,
            updated_at=excluded.updated_at
        """,
        (paper_id, workbook.name, title, doi, year, journal, json.dumps(paper_data, ensure_ascii=False), imported_at),
    )

    counts = {"records": 0, "review_queue": 0}
    for sheet, rows in sheet_rows.items():
        if sheet == "review_queue":
            for index, row in enumerate(rows, 1):
                review_key = str(pick(row, ID_COLUMNS["review_queue"]) or f"{paper_id}:review:{index}")
                status = initial_status(row)
                conn.execute(
                    """
                    INSERT INTO review_queue(
                        source_workbook, paper_id, review_key, target_sheet, target_record_ids,
                        priority, question, suggested_action, status, final_decision,
                        human_answer, notes, data_json, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(source_workbook, review_key) DO UPDATE SET
                        paper_id=excluded.paper_id,
                        target_sheet=excluded.target_sheet,
                        target_record_ids=excluded.target_record_ids,
                        priority=excluded.priority,
                        question=excluded.question,
                        suggested_action=excluded.suggested_action,
                        status=CASE
                            WHEN EXISTS(SELECT 1 FROM answers WHERE answers.review_item_id=review_queue.id)
                            THEN review_queue.status
                            ELSE excluded.status
                        END,
                        final_decision=CASE
                            WHEN EXISTS(SELECT 1 FROM answers WHERE answers.review_item_id=review_queue.id)
                            THEN review_queue.final_decision
                            ELSE excluded.final_decision
                        END,
                        human_answer=CASE
                            WHEN EXISTS(SELECT 1 FROM answers WHERE answers.review_item_id=review_queue.id)
                            THEN review_queue.human_answer
                            ELSE excluded.human_answer
                        END,
                        notes=CASE
                            WHEN EXISTS(SELECT 1 FROM answers WHERE answers.review_item_id=review_queue.id)
                            THEN review_queue.notes
                            ELSE excluded.notes
                        END,
                        data_json=excluded.data_json,
                        updated_at=excluded.updated_at
                    """,
                    (
                        workbook.name,
                        str(pick(row, ["paper_id"]) or paper_id),
                        review_key,
                        pick(row, ["target_sheet", "sheet"]),
                        pick(row, ["target_record_ids", "target_records", "record_ids"]),
                        pick(row, ["priority"]),
                        pick(row, ["question", "issue", "review_question"]),
                        pick(row, ["suggested_action", "suggestion"]),
                        status,
                        pick(row, ["final_decision", "human_decision"]),
                        pick(row, ["human_answer", "answer"]),
                        pick(row, ["notes", "note"]),
                        json.dumps(row, ensure_ascii=False),
                        imported_at,
                    ),
                )
                counts["review_queue"] += 1
            continue

        id_candidates = ID_COLUMNS.get(sheet, ["id"])
        for index, row in enumerate(rows, 1):
            row_paper_id = str(pick(row, ["paper_id"]) or paper_id)
            record_key = str(pick(row, id_candidates) or f"{row_paper_id}:{sheet}:{index}")
            conn.execute(
                """
                INSERT INTO records(source_workbook, paper_id, sheet, record_key, human_decision, confidence, data_json, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(source_workbook, sheet, record_key) DO UPDATE SET
                    paper_id=excluded.paper_id,
                    human_decision=excluded.human_decision,
                    confidence=excluded.confidence,
                    data_json=excluded.data_json,
                    updated_at=excluded.updated_at
                """,
                (
                    workbook.name,
                    row_paper_id,
                    sheet,
                    record_key,
                    pick(row, ["human_decision", "decision"]),
                    pick(row, ["confidence"]),
                    json.dumps(row, ensure_ascii=False),
                    imported_at,
                ),
            )
            counts["records"] += 1

    conn.execute(
        "INSERT INTO events(event_type, payload_json, created_at) VALUES (?, ?, ?)",
        ("import_workbook", json.dumps({"workbook": workbook.name, **counts}, ensure_ascii=False), imported_at),
    )
    return counts


def import_directory(db_path: Path = DEFAULT_DB, annotation_dir: Path = DEFAULT_ANNOTATION_DIR) -> dict[str, int]:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    workbooks = sorted(annotation_dir.glob("*.xlsx"))
    totals = {"workbooks": 0, "records": 0, "review_queue": 0}
    with connect(db_path) as conn:
        create_schema(conn)
        for workbook in workbooks:
            counts = import_workbook(conn, workbook)
            totals["workbooks"] += 1
            totals["records"] += counts["records"]
            totals["review_queue"] += counts["review_queue"]
        conn.commit()
    return totals


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--annotation-dir", type=Path, default=DEFAULT_ANNOTATION_DIR)
    args = parser.parse_args()
    totals = import_directory(args.db, args.annotation_dir)
    print(json.dumps({"database": str(args.db), **totals}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
